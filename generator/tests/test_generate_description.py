import asyncio
import io
import json
import unittest

from openpyxl import Workbook, load_workbook

from app.use_cases.generate_description import (
    MAX_MARKUP_CHARS,
    MAX_ROWS,
    OUTPUT_COLUMNS,
    DescriptionGenerationUseCase,
    InvalidDescriptionInput,
)


GENERATED_FIELDS = {
    "Описание": "Краткое описание.",
    "Какое действующее вещество?": "Вещество 10 мг.",
    "От чего помогает?": "От симптомов.",
    "Как принимать?": "По инструкции.",
    "Нужен рецепт?": "По рецепту",
    "С чем сочетается?": "Нет данных.",
    "Какие побочные эффекты?": "Редкие реакции.",
    "Как хранить?": "При температуре не выше 25 °C.",
}


class FakeResponse:
    def __init__(self, content: str) -> None:
        self.content = content


class FakeLlm:
    def __init__(self) -> None:
        self.calls = []

    async def ainvoke(self, messages):
        self.calls.append(messages)
        return FakeResponse(json.dumps(GENERATED_FIELDS, ensure_ascii=False))


def workbook_bytes(rows: list[tuple[object, object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class DescriptionGenerationUseCaseTests(unittest.IsolatedAsyncioTestCase):
    async def test_generates_xlsx_from_body(self):
        llm = FakeLlm()
        use_case = DescriptionGenerationUseCase(llm=llm)

        result = await use_case.execute(item_id="72128", raw_description="сырая разметка")

        sheet = load_workbook(io.BytesIO(result), read_only=True).active
        self.assertEqual([cell.value for cell in sheet[1]], OUTPUT_COLUMNS)
        self.assertEqual(sheet.cell(2, 1).value, "72128")
        self.assertEqual(sheet.cell(2, 2).value, GENERATED_FIELDS["Описание"])
        self.assertEqual(len(llm.calls), 1)
        self.assertEqual(llm.calls[0][1].content, "сырая разметка")

    async def test_generates_rows_from_headerless_example_workbook(self):
        llm = FakeLlm()
        use_case = DescriptionGenerationUseCase(llm=llm)
        source = workbook_bytes([(72128, "raw one"), (14820, "raw two")])

        result = await use_case.execute(file_bytes=source, filename="input.xlsx")

        sheet = load_workbook(io.BytesIO(result), read_only=True).active
        self.assertEqual(sheet.max_row, 3)
        self.assertEqual(sheet.cell(2, 1).value, 72128)
        self.assertEqual(sheet.cell(3, 1).value, 14820)
        self.assertEqual([call[1].content for call in llm.calls], ["raw one", "raw two"])

    async def test_accepts_workbook_with_named_columns(self):
        use_case = DescriptionGenerationUseCase(llm=FakeLlm())
        source = workbook_bytes([("id", "разметка сырая"), (7, "raw")])

        result = await use_case.execute(file_bytes=source, filename="input.xlsx")

        sheet = load_workbook(io.BytesIO(result), read_only=True).active
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.cell(2, 1).value, 7)

    async def test_rejects_more_or_less_than_one_input_variant(self):
        use_case = DescriptionGenerationUseCase(llm=FakeLlm())

        with self.assertRaises(InvalidDescriptionInput):
            await use_case.execute()
        with self.assertRaises(InvalidDescriptionInput):
            await use_case.execute(
                file_bytes=workbook_bytes([(1, "raw")]),
                filename="input.xlsx",
                item_id="1",
                raw_description="raw",
            )

    async def test_records_incomplete_llm_json(self):
        class IncompleteLlm:
            async def ainvoke(self, messages):
                return FakeResponse('{"Описание": "Только одно поле"}')

        use_case = DescriptionGenerationUseCase(
            llm=IncompleteLlm(),
            retry_base_delay_seconds=0,
        )

        result = await use_case.execute(item_id="1", raw_description="raw")

        sheet = load_workbook(io.BytesIO(result), read_only=True).active
        self.assertTrue(sheet.cell(2, 2).value.startswith("Ошибка генерации: "))
        self.assertEqual(sheet.cell(2, 3).value, "Нет данных.")

    async def test_neutralizes_formula_like_cell_values(self):
        fields = {**GENERATED_FIELDS, "Описание": '=HYPERLINK("https://example.test")'}

        class FormulaLlm:
            async def ainvoke(self, messages):
                return FakeResponse(json.dumps(fields, ensure_ascii=False))

        result = await DescriptionGenerationUseCase(llm=FormulaLlm()).execute(
            item_id="=1+1",
            raw_description="raw",
        )

        sheet = load_workbook(io.BytesIO(result), read_only=True).active
        self.assertEqual(sheet.cell(2, 1).data_type, "s")
        self.assertEqual(sheet.cell(2, 1).value, "'=1+1")
        self.assertEqual(sheet.cell(2, 2).data_type, "s")
        self.assertTrue(sheet.cell(2, 2).value.startswith("'="))

    async def test_rejects_empty_id_and_oversized_markup(self):
        use_case = DescriptionGenerationUseCase(llm=FakeLlm())

        with self.assertRaises(InvalidDescriptionInput):
            await use_case.execute(item_id="  ", raw_description="raw")
        with self.assertRaises(InvalidDescriptionInput):
            await use_case.execute(item_id="1", raw_description="x" * (MAX_MARKUP_CHARS + 1))

    async def test_rejects_workbook_with_too_many_rows(self):
        source = workbook_bytes([(index, "raw") for index in range(MAX_ROWS + 1)])

        with self.assertRaises(InvalidDescriptionInput):
            await DescriptionGenerationUseCase(llm=FakeLlm()).execute(
                file_bytes=source,
                filename="input.xlsx",
            )

    async def test_limits_concurrency_and_preserves_input_order(self):
        class ConcurrentLlm:
            def __init__(self) -> None:
                self.active_calls = 0
                self.peak_active_calls = 0

            async def ainvoke(self, messages):
                self.active_calls += 1
                self.peak_active_calls = max(self.peak_active_calls, self.active_calls)
                await asyncio.sleep(0.01)
                self.active_calls -= 1
                fields = {**GENERATED_FIELDS, "Описание": messages[1].content}
                return FakeResponse(json.dumps(fields, ensure_ascii=False))

        llm = ConcurrentLlm()
        source = workbook_bytes([(3, "third"), (1, "first"), (2, "second")])
        use_case = DescriptionGenerationUseCase(llm=llm, max_concurrency=2)

        result = await use_case.execute(file_bytes=source, filename="input.xlsx")

        sheet = load_workbook(io.BytesIO(result), read_only=True).active
        self.assertEqual(llm.peak_active_calls, 2)
        self.assertEqual(
            [(sheet.cell(row, 1).value, sheet.cell(row, 2).value) for row in range(2, 5)],
            [(3, "third"), (1, "first"), (2, "second")],
        )

    async def test_retries_invalid_json_and_transient_llm_error(self):
        class FlakyLlm:
            def __init__(self) -> None:
                self.attempts: dict[str, int] = {}

            async def ainvoke(self, messages):
                markup = messages[1].content
                attempt = self.attempts.get(markup, 0) + 1
                self.attempts[markup] = attempt
                if markup == "invalid-json" and attempt == 1:
                    return FakeResponse("not json")
                if markup == "temporary-error" and attempt == 1:
                    raise TimeoutError("temporary")
                return FakeResponse(json.dumps(GENERATED_FIELDS, ensure_ascii=False))

        llm = FlakyLlm()
        source = workbook_bytes([(1, "invalid-json"), (2, "temporary-error")])
        use_case = DescriptionGenerationUseCase(llm=llm, retry_base_delay_seconds=0)

        result = await use_case.execute(file_bytes=source, filename="input.xlsx")

        sheet = load_workbook(io.BytesIO(result), read_only=True).active
        self.assertEqual(sheet.max_row, 3)
        self.assertEqual(llm.attempts, {"invalid-json": 2, "temporary-error": 2})

    async def test_records_row_error_without_failing_other_rows(self):
        class PartiallyFailingLlm:
            def __init__(self) -> None:
                self.attempts: dict[str, int] = {}

            async def ainvoke(self, messages):
                markup = messages[1].content
                self.attempts[markup] = self.attempts.get(markup, 0) + 1
                if markup == "broken":
                    return FakeResponse('{"Описание":')
                fields = {**GENERATED_FIELDS, "Описание": markup}
                return FakeResponse(json.dumps(fields, ensure_ascii=False))

        llm = PartiallyFailingLlm()
        source = workbook_bytes([(10, "ok-before"), (20, "broken"), (30, "ok-after")])
        use_case = DescriptionGenerationUseCase(llm=llm, retry_base_delay_seconds=0)

        result = await use_case.execute(file_bytes=source, filename="input.xlsx")

        sheet = load_workbook(io.BytesIO(result), read_only=True).active
        self.assertEqual([sheet.cell(row, 1).value for row in range(2, 5)], [10, 20, 30])
        self.assertEqual(sheet.cell(2, 2).value, "ok-before")
        self.assertTrue(sheet.cell(3, 2).value.startswith("Ошибка генерации: "))
        self.assertEqual(sheet.cell(3, 3).value, "Нет данных.")
        self.assertEqual(sheet.cell(4, 2).value, "ok-after")
        self.assertEqual(llm.attempts["broken"], 3)


if __name__ == "__main__":
    unittest.main()
