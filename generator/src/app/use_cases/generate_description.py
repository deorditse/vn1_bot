import asyncio
import io
import json
from collections.abc import Sequence
from itertools import islice
from pathlib import Path
from typing import Any

import xlrd
from fastapi import Request
from langchain_core.messages import HumanMessage, SystemMessage
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from starlette.datastructures import UploadFile as StarletteUploadFile

from app.api.schemas.generate import GenerateDescriptionBody
from app.policies.policies_loader import load_prompt_text
from infrastructure.llm.llm import LLMService

OUTPUT_FIELDS = [
    "Описание",
    "Какое действующее вещество?",
    "От чего помогает?",
    "Как принимать?",
    "Нужен рецепт?",
    "С чем сочетается?",
    "Какие побочные эффекты?",
    "Как хранить?",
]
OUTPUT_COLUMNS = ["id", *OUTPUT_FIELDS]
MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_ROWS = 100
MAX_MARKUP_CHARS = 100_000
MAX_TOTAL_MARKUP_CHARS = 1_000_000
MAX_REQUEST_BYTES = MAX_FILE_BYTES + 64 * 1024
MAX_JSON_BYTES = MAX_MARKUP_CHARS + 2048
DEFAULT_MAX_CONCURRENCY = 4
DEFAULT_MAX_ATTEMPTS = 3
DEFAULT_RETRY_BASE_DELAY_SECONDS = 0.5
NO_DATA = "Нет данных."
GENERATION_ERROR_PREFIX = "Ошибка генерации: "
_ID_HEADERS = {"id", "идентификатор"}
_RAW_HEADERS = {
    "разметка сырая",
    "сырая разметка",
    "raw_description",
    "raw description",
    "raw_markup",
    "raw markup",
}


class InvalidDescriptionInput(ValueError):
    pass


class DescriptionGenerationUseCase:
    ai_information_prompt = load_prompt_text("generation/from_markup.md")

    def __init__(
        self,
        llm: Any | None = None,
        *,
        max_concurrency: int = DEFAULT_MAX_CONCURRENCY,
        max_attempts: int = DEFAULT_MAX_ATTEMPTS,
        retry_base_delay_seconds: float = DEFAULT_RETRY_BASE_DELAY_SECONDS,
    ) -> None:
        if max_concurrency < 1:
            raise ValueError("max_concurrency должен быть больше нуля")
        if max_attempts < 1:
            raise ValueError("max_attempts должен быть больше нуля")
        if retry_base_delay_seconds < 0:
            raise ValueError("retry_base_delay_seconds не может быть отрицательным")

        self._llm = llm or LLMService().openai()
        self._max_concurrency = max_concurrency
        self._max_attempts = max_attempts
        self._retry_base_delay_seconds = retry_base_delay_seconds

    async def execute_request(self, request: Request) -> bytes:
        self._validate_content_length(request)
        content_type = request.headers.get("content-type", "").lower()

        if content_type.startswith("multipart/form-data"):
            async with request.form(max_files=1, max_fields=2, max_part_size=64 * 1024) as form:
                items = list(form.multi_items())
                if len(items) != 1 or items[0][0] != "file" or not isinstance(items[0][1], StarletteUploadFile):
                    raise InvalidDescriptionInput("Передайте ровно один XLS/XLSX-файл в поле file")
                file = items[0][1]
                return await self.execute(
                    file_bytes=await self._read_upload_limited(file),
                    filename=file.filename or "input.xlsx",
                )

        if content_type.startswith("application/json"):
            try:
                body = await self._read_request_body_limited(request, MAX_JSON_BYTES)
                payload = GenerateDescriptionBody.model_validate(json.loads(body))
            except (json.JSONDecodeError, UnicodeDecodeError, ValidationError) as error:
                raise InvalidDescriptionInput("Некорректный JSON body") from error
            return await self.execute(
                item_id=payload.id,
                raw_description=payload.raw_description,
            )

        raise InvalidDescriptionInput(
            "Используйте multipart/form-data с файлом или application/json"
        )

    async def execute(
        self,
        *,
        file_bytes: bytes | None = None,
        filename: str | None = None,
        item_id: str | int | None = None,
        raw_description: str | None = None,
    ) -> bytes:
        has_file = file_bytes is not None
        has_body = item_id is not None or raw_description is not None
        if has_file == has_body:
            raise InvalidDescriptionInput(
                "Передайте ровно один источник: XLS/XLSX-файл или поля id и raw_description"
            )

        if has_file:
            if len(file_bytes or b"") > MAX_FILE_BYTES:
                raise InvalidDescriptionInput("Размер файла не должен превышать 10 МБ")
            rows = self._read_rows(file_bytes or b"", filename or "input.xlsx")
        else:
            if item_id is None or not str(item_id).strip() or not str(raw_description or "").strip():
                raise InvalidDescriptionInput("Поля id и raw_description обязательны")
            rows = [(item_id, raw_description or "")]

        if len(rows) > MAX_ROWS:
            raise InvalidDescriptionInput(f"Таблица должна содержать не более {MAX_ROWS} строк")
        total_markup_chars = 0
        for row_id, markup in rows:
            if len(str(row_id)) > 128:
                raise InvalidDescriptionInput("id не должен превышать 128 символов")
            markup_length = len(markup)
            if markup_length > MAX_MARKUP_CHARS:
                raise InvalidDescriptionInput(
                    f"Сырая разметка одной строки не должна превышать {MAX_MARKUP_CHARS} символов"
                )
            total_markup_chars += markup_length
        if total_markup_chars > MAX_TOTAL_MARKUP_CHARS:
            raise InvalidDescriptionInput("Общий объем сырой разметки слишком большой")

        # Семафор пропускает к LLM не больше заданного количества строк одновременно.
        semaphore = asyncio.Semaphore(self._max_concurrency)

        async def generate(
            index: int,
            row_id: Any,
            markup: str,
        ) -> tuple[int, Any, dict[str, str]]:
            # Индекс нужен, чтобы завершившиеся в разное время запросы не изменили порядок строк.
            return (
                index,
                row_id,
                await self._generate_fields(markup, semaphore),
            )

        # gather конкурентно запускает обработку всех строк и ждет завершения каждой из них.
        indexed_rows = await asyncio.gather(
            *(generate(index, row_id, markup) for index, (row_id, markup) in enumerate(rows))
        )
        indexed_rows.sort(key=lambda row: row[0])

        return self._build_workbook([(row_id, fields) for _, row_id, fields in indexed_rows])

    async def _generate_fields(
        self,
        markup: str,
        semaphore: asyncio.Semaphore,
    ) -> dict[str, str]:
        last_error: Exception | None = None

        for attempt in range(self._max_attempts):
            try:
                # Слот семафора занят только во время запроса; пауза перед повтором его не блокирует.
                async with semaphore:
                    response = await self._llm.ainvoke(
                        [
                            SystemMessage(content=self.ai_information_prompt.strip()),
                            HumanMessage(content=markup),
                        ]
                    )
                return self._parse_response(response.content)
            except Exception as error:
                last_error = error
                if attempt + 1 < self._max_attempts:
                    # Каждый следующий повтор откладывается вдвое дольше предыдущего.
                    await asyncio.sleep(self._retry_base_delay_seconds * (2**attempt))

        return self._error_fields(last_error)

    @staticmethod
    def _error_fields(error: Exception | None) -> dict[str, str]:
        if isinstance(error, InvalidDescriptionInput):
            reason = str(error)
        elif error is not None:
            reason = type(error).__name__
        else:
            reason = "неизвестная ошибка"

        return {
            field: f"{GENERATION_ERROR_PREFIX}{reason}" if field == "Описание" else NO_DATA
            for field in OUTPUT_FIELDS
        }

    @classmethod
    def _read_rows(cls, file_bytes: bytes, filename: str) -> list[tuple[Any, str]]:
        suffix = Path(filename).suffix.lower()
        if suffix == ".xls":
            rows = cls._read_xls(file_bytes)
        elif suffix == ".xlsx":
            rows = cls._read_xlsx(file_bytes)
        else:
            raise InvalidDescriptionInput("Поддерживаются только файлы .xls и .xlsx")

        if not rows:
            raise InvalidDescriptionInput("Таблица не содержит данных")

        first_id = cls._normalize_header(rows[0][0])
        first_raw = cls._normalize_header(rows[0][1])
        if first_id in _ID_HEADERS and first_raw in _RAW_HEADERS:
            rows = rows[1:]

        result: list[tuple[Any, str]] = []
        for row_number, (row_id, markup) in enumerate(rows, start=1):
            if row_id is None and markup is None:
                continue
            if row_id is None or not str(markup or "").strip():
                raise InvalidDescriptionInput(
                    f"Строка {row_number}: должны быть заполнены id и сырая разметка"
                )
            result.append((row_id, str(markup)))

        if not result:
            raise InvalidDescriptionInput("Таблица не содержит строк для генерации")
        return result

    @staticmethod
    def _read_xlsx(file_bytes: bytes) -> list[tuple[Any, Any]]:
        try:
            workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            try:
                sheet = workbook.active
                if sheet is None:
                    return []
                rows = sheet.iter_rows(min_col=1, max_col=2, values_only=True)
                return [(row[0], row[1]) for row in islice(rows, MAX_ROWS + 2)]
            finally:
                workbook.close()
        except Exception as error:
            raise InvalidDescriptionInput("Не удалось прочитать XLSX-файл") from error

    @staticmethod
    def _read_xls(file_bytes: bytes) -> list[tuple[Any, Any]]:
        workbook = None
        try:
            workbook = xlrd.open_workbook(file_contents=file_bytes)
            sheet = workbook.sheet_by_index(0)
            return [
                (sheet.cell_value(index, 0), sheet.cell_value(index, 1))
                for index in range(min(sheet.nrows, MAX_ROWS + 2))
            ]
        except Exception as error:
            raise InvalidDescriptionInput("Не удалось прочитать XLS-файл") from error
        finally:
            if workbook is not None:
                workbook.release_resources()

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _parse_response(content: Any) -> dict[str, str]:
        if not isinstance(content, str):
            raise InvalidDescriptionInput("LLM вернула ответ в неподдерживаемом формате")

        raw_json = content.strip()
        if raw_json.startswith("```"):
            lines = raw_json.splitlines()
            raw_json = "\n".join(lines[1:-1]).strip()

        try:
            payload = json.loads(raw_json)
        except json.JSONDecodeError as error:
            raise InvalidDescriptionInput("LLM вернула некорректный JSON") from error

        if not isinstance(payload, dict):
            raise InvalidDescriptionInput("LLM должна вернуть JSON-объект")

        missing = [field for field in OUTPUT_FIELDS if not str(payload.get(field, "")).strip()]
        if missing:
            raise InvalidDescriptionInput(f"В ответе LLM отсутствуют поля: {', '.join(missing)}")
        return {field: str(payload[field]).strip() for field in OUTPUT_FIELDS}

    @staticmethod
    def _build_workbook(rows: Sequence[tuple[Any, dict[str, str]]]) -> bytes:
        workbook = Workbook()
        buffer = io.BytesIO()
        try:
            sheet = workbook.active
            if sheet is None:
                raise RuntimeError("Не удалось создать лист XLSX")
            sheet.title = "Лист1"
            sheet.append(OUTPUT_COLUMNS)

            for row_id, fields in rows:
                sheet.append(
                    [
                        DescriptionGenerationUseCase._safe_cell_value(row_id),
                        *(DescriptionGenerationUseCase._safe_cell_value(fields[field]) for field in OUTPUT_FIELDS),
                    ]
                )

            header_fill = PatternFill("solid", fgColor="E7E6E6")
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            widths = [14, 42, 30, 42, 42, 18, 42, 42, 32]
            for index, width in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(index)].width = width
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

            workbook.save(buffer)
            return buffer.getvalue()
        finally:
            workbook.close()
            buffer.close()

    @staticmethod
    def _validate_content_length(request: Request) -> None:
        value = request.headers.get("content-length")
        if not value:
            return
        try:
            content_length = int(value)
        except ValueError as error:
            raise InvalidDescriptionInput("Некорректный Content-Length") from error
        if content_length > MAX_REQUEST_BYTES:
            raise InvalidDescriptionInput("Размер запроса не должен превышать 10 МБ")

    @staticmethod
    async def _read_upload_limited(file: StarletteUploadFile) -> bytes:
        chunks: list[bytes] = []
        size = 0
        while chunk := await file.read(64 * 1024):
            size += len(chunk)
            if size > MAX_FILE_BYTES:
                raise InvalidDescriptionInput("Размер файла не должен превышать 10 МБ")
            chunks.append(chunk)
        return b"".join(chunks)

    @staticmethod
    async def _read_request_body_limited(request: Request, limit: int) -> str:
        chunks: list[bytes] = []
        size = 0
        async for chunk in request.stream():
            size += len(chunk)
            if size > limit:
                raise InvalidDescriptionInput("JSON body слишком большой")
            chunks.append(chunk)
        return b"".join(chunks).decode("utf-8")

    @staticmethod
    def _safe_cell_value(value: Any) -> Any:
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value
